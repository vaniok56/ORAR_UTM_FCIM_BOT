import numpy as np 
import openpyxl # excel read library
import datetime
import pytz
import requests
import handlers.db as db

import time
from collections import defaultdict
last_command_time = defaultdict(float)
messages_per_minute = defaultdict(list)
COMMAND_COOLDOWN = 1 # seconds
MAX_MESSAGES_PER_MINUTE = 5 # messages

moldova_tz = pytz.timezone('Europe/Chisinau')
time_zone = pytz.timezone('Europe/Chisinau')
current_year = 26
#logs
import logging
class ColoredFormatter(logging.Formatter):
    COLORS = {
        'DEBUG': '\033[94m',  # Blue
        'INFO': '\033[92m',   # Green
        'WARNING': '\033[93m', # Yellow
        'ERROR': '\033[91m',   # Red
        'CRITICAL': '\033[91m\033[1m', # Bold Red
        'RESET': '\033[0m'    # Reset color
    }

    def format(self, record):
        # Format without the levelname, which we'll add separately with color
        formatter = logging.Formatter(
            '%(asctime)s.%(msecs)03d | %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        
        # Get the log message without the level
        log_message = formatter.format(record)
        
        # Add the colored level tag
        colored_level = f"{self.COLORS.get(record.levelname, self.COLORS['RESET'])}[{record.levelname}]{self.COLORS['RESET']}"
        
        # Insert the colored level at the position after the timestamp
        parts = log_message.split(' | ', 1)
        if len(parts) == 2:
            return f"{parts[0]} | {colored_level} {parts[1]}"
        return log_message

# Create formatters - one with color for console, one without for file
file_formatter = logging.Formatter(
    '%(asctime)s.%(msecs)03d | [%(levelname)s] %(message)s',
    datefmt='%H:%M:%S'
)
console_formatter = ColoredFormatter(
    '%(asctime)s.%(msecs)03d | [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

# Create handlers
date = datetime.datetime.now(time_zone).strftime("%d_%m_%y")
log_dir = "logs"
import os
# Ensure logs directory exists
os.makedirs(log_dir, exist_ok=True)
try:
    file_handler = logging.FileHandler(f"{log_dir}/orarbot_{date}.log")
    file_handler.setFormatter(file_formatter)
except Exception as e:
    print(f"Error creating log file: {e}")
    # Fallback to a null handler if file can't be created
    file_handler = logging.NullHandler()

console_handler = logging.StreamHandler()
console_handler.setFormatter(console_formatter)

# Configure the root logger
logging.basicConfig(
    level=logging.INFO,
    handlers=[file_handler, console_handler]
)

logging.Formatter.converter = lambda *args: \
    datetime.datetime.now(time_zone).timetuple()

is_even = (datetime.datetime.now(moldova_tz)).isocalendar().week % 2

#hours
hours =   [
    ["8.00-9.30"],
    ["9.45-11.15"],
    ["11.30-13.00"],
    ["13.30-15.00"],
    ["15.15-16.45"],
    ["17.00-18.30"],
    ["18.45-20.15"]
]

#week days
week_days = {
    0 : "Luni",
    1 : "Marţi",
    2 : "Miercuri",
    3 : "Joi",
    4 : "Vineri",
    5 : "Sâmbătă",
    6 : "Duminica"
}

all_groups_range = range(3, 40)
schedule_column_start = 3

cur_group = "TI-241" #initialise current group

#cache
day_row_start_cache = {}  #cache for row start by day name and schedule
daily_schedule_cache = {}  #cache for daily schedules
cell_value_cache = {}  #cache for cell values
weekly_schedule_cache = {} #cache for weekly schedules
next_course_cache = {} #cache for next course
orele_cache = {} #cache for hours
schedule_groups_cache = {} #cache for schedule and groups

def send_logs(message, type):
    if type =='info':
        logging.info(message)
    elif type =='warning':
        logging.warning(message)
    elif type =='error':
        logging.error(message)
    elif type =='critical':
        logging.critical(message)
    else: 
        logging.info(message)

#open the excel files
for i in range(1, 5):
    try:
        globals()[f"schedule{i}"] = openpyxl.load_workbook(f'orar{i}.xlsx', data_only=True)["Table 2"]
    except Exception as e:
        send_logs(f"Error loading orar{i}.xlsx: {e}", 'error')
        globals()[f"schedule{i}"] = openpyxl.Workbook()
        globals()[f"schedule{i}"].create_sheet("Table 2")

#group lists
for i in range(1, 5):
    try:
        globals()[f"groups{i}"] = [globals()[f"schedule{i}"].cell(row=1,column=j).value for j in all_groups_range if globals()[f"schedule{i}"].cell(row=1,column=j).value]
        send_logs(f"Extracted {len(globals()[f"groups{i}"])} groups from schedule{i}", 'info')
    except Exception as e:
        send_logs(f"Error extracting groups from schedule{i}: {e}", 'error')
        globals()[f"groups{i}"] = []

def get_schedule_and_groups(cur_group):
    if cur_group in schedule_groups_cache:
        #send_logs(f"Cache hit get_schedule_and_groups for {cur_group}", 'info')
        return schedule_groups_cache[cur_group]

    group_year = int(cur_group[-3:-1])
    sch_nr = current_year - group_year
    if sch_nr >= 1 and sch_nr <= 4:
        result = globals()[f"schedule{sch_nr}"], globals()[f"groups{sch_nr}"]
    else:
        raise ValueError("Invalid group number")
    #cache
    schedule_groups_cache[cur_group] = result
    return result

#get value from a cell even if it's a merged cell
merged_cell_ranges = {}  # cache merged cell ranges
def get_merged_cell_ranges(sheet):
    global merged_cell_ranges
    if sheet not in merged_cell_ranges:
        merged_cell_ranges[sheet] = {(r.min_row, r.min_col): r for r in sheet.merged_cells.ranges}
    return merged_cell_ranges[sheet]
def getMergedCellVal(sheet, cell):
    cell_key = (id(sheet), cell.row, cell.column)
    if cell_key in cell_value_cache:
        #send_logs(f"Cache hit getMergedCellVal for {cell_key}", 'info')
        return cell_value_cache[cell_key]
    
    merged_ranges = get_merged_cell_ranges(sheet)
    row, col = cell.row, cell.column
    
    for (r_min, c_min), rng in merged_ranges.items():
        if r_min <= row <= rng.max_row and c_min <= col <= rng.max_col:
            value = sheet.cell(rng.min_row, rng.min_col).value
            # Cache the value
            cell_value_cache[cell_key] = value
            return value
    
    value = cell.value
    cell_value_cache[cell_key] = value
    return value

def button_grid(buttons, butoane_rand):
    grid = []
    row = []
    for button in buttons:
        if button.text == "Back":
            if row:
                grid.append(row)
            row = [button]
        else:
            row.append(button)
        if len(row) != butoane_rand:
            continue
        grid.append(row)
        row = []
    if row:
        grid.append(row)
    return grid

#get daily schedule
def print_day(week_day, cur_group, is_even, subgrupa):
    schedule, groups = get_schedule_and_groups(cur_group)[0:2]
    col_gr = groups.index(cur_group) + schedule_column_start  # column with the selected group
    return print_daily(schedule, is_even, col_gr, week_day, subgrupa)
     
#extract daily schedule
def print_daily(schedule, is_even, col_gr, week_day, subgrupa):
    #cache key
    cache_key = (id(schedule), is_even, col_gr, week_day, subgrupa)
    
    #find daily schedule in cache
    if cache_key in daily_schedule_cache:
        #send_logs(f"Cache hit schedule for {cache_key}", 'info')
        return daily_schedule_cache[cache_key]
    
    #subgrupa - 0/1/2
    try:
        subgrupa = int(subgrupa)
    except (ValueError, TypeError):
        subgrupa = 0
        
    #if is even, change the subgrupa
    if is_even and subgrupa != 0:
        subgrupa = 3 - subgrupa

    day_sch = []
    seen = set() 
    day_name = week_days[week_day]
    
    #find row start in cache
    schedule_day_key = (id(schedule), day_name, is_even)
    if schedule_day_key in day_row_start_cache:
        #send_logs(f"Cache hit row_start for {schedule_day_key}", 'info')
        row_start = day_row_start_cache[schedule_day_key]
    else:
        for i in range(1, 84):
            if day_name != getMergedCellVal(schedule, schedule.cell(row=i, column=1)):
                continue
            row_start = i
            #cache row start
            day_row_start_cache[schedule_day_key] = row_start
            break
        else:
            #cache empty result
            daily_schedule_cache[cache_key] = ""
            return ""
    
    
    orele_key = (id(schedule), row_start)
    if orele_key in orele_cache:
        #send_logs(f"Cache hit orele for {orele_key}", 'info')
        orele = orele_cache[orele_key]
    else:
        orele = {i: getMergedCellVal(schedule, schedule.cell(row=i, column=2)) 
                for i in range(row_start, row_start + 14)}
        orele_cache[orele_key] = orele
    #extract the daily schedule
    for i in range(row_start, row_start + 13):
        if orele[i] in seen:
            continue
        if is_even and orele.get(i + 1) == orele[i]:
            cell_value = getMergedCellVal(schedule, schedule.cell(row=i + 1, column=col_gr))
        else:
            cell_value = getMergedCellVal(schedule, schedule.cell(row=i, column=col_gr))
        day_sch.append(cell_value)
        seen.add(orele[i])

    processed_courses = []

    for i, course in enumerate(day_sch):
        if course is None or course == "":
            continue

        if subgrupa != 0:
            count_05 = course.count("0.5") + course.count("0,5")
            if count_05 == 2:
                if subgrupa == 1:
                    try:
                        course = course.split("\n2)")[0]
                    except:
                        course = course
                else:
                    course = "2)" + course.split("\n2)")[1]
            elif count_05 == 1:
                if subgrupa == 2:
                    course = ""
        
        if course:
            processed_courses.append(f"\nPerechea: #{i + 1}\n<b>{course}</b>\nOra : {hours[i][0].replace('.', ':')}\n")
            
    # Join the properly formatted strings
    result = "".join(processed_courses)
    daily_schedule_cache[cache_key] = result
    return result

def print_next_course(week_day, cur_group, is_even, course_index, subgrupa):
    cache_key = (week_day, cur_group, is_even, course_index, subgrupa)
    if cache_key in next_course_cache:
        #send_logs(f"Cache hit next_course for {cache_key}", 'info')
        return next_course_cache[cache_key]
    
    #get daily schedule
    daily = print_day(week_day, cur_group, is_even, subgrupa)
    if not daily:
        next_course_cache[cache_key] = ""
        return ""
    
    courses = daily.split("Perechea: #")
    for i in range(1, len(courses)):
        if int(courses[i][0]) != course_index:
            continue
        course = courses[i]
        parts = course.split("Ora : ")
        course_name = parts[0][1:]  # Skip the index digit
        course_time = parts[1]
        result = f"<b>{course_name}</b>Ora: {course_time}"
        next_course_cache[cache_key] = result
        return result
    
    next_course_cache[cache_key] = ""
    return ""

#get weekly schedule
def print_sapt(is_even, cur_group, subgrupa):
    cache_key = (cur_group, is_even, subgrupa)
    if cache_key in weekly_schedule_cache:
        #send_logs(f"Cache hit print_sapt for {cache_key}", 'info')
        return weekly_schedule_cache[cache_key]
    schedule, groups = get_schedule_and_groups(cur_group)[0:2]
    col_gr = groups.index(cur_group) + 3 #column with the selected group

    week_sch = ""
    for j in range(1, 7):
        daily = print_daily(schedule, is_even, col_gr, j-1, subgrupa)
        #do not print an empty weekday
        if str(daily) == "":
            continue
        week_sch += "\n\n&emsp;&emsp;&emsp;&emsp;<b>" + week_days[j-1] + ":</b>" + "\n" + daily
    weekly_schedule_cache[cache_key] = week_sch
    return week_sch

def get_next_course_time():
    current_time = datetime.datetime.now(moldova_tz).time()
    current_time = datetime.datetime.strptime(str(current_time)[:-7], "%H:%M:%S")
    
    #find next course index
    course_index = 0
    for i, hour in enumerate(hours):
        course_time = datetime.datetime.strptime(hour[0].split("-")[0], "%H.%M")
        if (course_time - datetime.timedelta(minutes=15)).time() > current_time.time():
            course_index = i
            break
    
    #15 min before the next course
    time_before_course = course_time - datetime.timedelta(minutes=15)
    
    return current_time, course_index + 1, time_before_course

def is_rate_limited(user_id):
    if user_id == 500303890:
        return False
    
    try:
        if db.locate_field("U"+str(user_id), "ban") == 1:
            return True
    except:
        pass

    current_time = time.time()
    minute_ago = current_time - 60

    messages = messages_per_minute.get(user_id, [])

    # Remove messages older than a minute
    if messages and messages[0] < minute_ago:
        messages = [t for t in messages if t > minute_ago]
    
    # Add current message timestamp
    messages.append(current_time)
    messages_per_minute[user_id] = messages
    
    # Apply 1-second cooldown only if more than 5 messages in the last minute
    if len(messages) > MAX_MESSAGES_PER_MINUTE:
        if current_time - last_command_time.get(user_id, 0) < COMMAND_COOLDOWN:
            return True
    
    last_command_time[user_id] = current_time
    return False

def format_id(user_id):
    return f"U{user_id}"

def get_version():    
    try:
        response = requests.get(
            "https://api.github.com/repos/vaniok56/ORAR_UTM_FCIM_BOT/commits",
            headers={"Accept": "application/vnd.github.v3+json"},
            timeout=5
        )
        
        if response.status_code == 200:
            latest_commit = response.json()[0]
            
            commit_date = latest_commit['commit']['author']['date']
            date_obj = datetime.datetime.strptime(commit_date, "%Y-%m-%dT%H:%M:%SZ")
            formatted_date = date_obj.strftime("%d-%m-%Y")
            
            commit_message = latest_commit['commit']['message']
            version_match = commit_message.split("Version ")[1].split(" ")[0] if "Version " in commit_message else None
            
            return version_match, formatted_date
            
    except Exception as e:
        send_logs(f"Error fetching version from GitHub: {e}", 'error')
        return "0.11.0", "01-09-2025"

def extract_specs(groups):
    specs = {}
    spec_order = []  # To preserve original order
    for group in groups:
        spec = group[:-4]  # Extract specialty part (e.g., "TI" from "TI-241")
        if spec not in specs:
            specs[spec] = []
            spec_order.append(spec)
        specs[spec].append(group)
    return specs, spec_order

def write_groups_to_json():
    years = {
        b"ye1": "  1  ",
        b"ye2": "  2  ",
        b"ye3": "  3  ",
        b"ye4": "  4  "
    }
    
    specialties = {}
    for year_num in ["1", "2", "3", "4"]:
        groups_var = globals()[f"groups{year_num}"]
        specs, spec_order = extract_specs(groups_var)
        
        specialties[year_num] = {}
        for spec in spec_order:  # Use ordered list
            key = f"{spec.lower()}{year_num}".encode()
            specialties[year_num][key] = f"  {spec}  "
    
    group_list = {}
    for year_num in ["1", "2", "3", "4"]:
        groups_var = globals()[f"groups{year_num}"]
        specs, spec_order = extract_specs(groups_var)
        
        group_list[year_num] = {}
        for spec in spec_order:  # Use ordered list
            spec_key = f"{spec}{year_num}"
            group_list[year_num][spec_key] = {}
            
            for group in groups_var:
                if not group.startswith(spec):
                    continue
                key = group.lower().encode()
                group_list[year_num][spec_key][key] = f"  {group}  "
    
    # Write to a Python file
    with open("dynamic_group_lists.py", "w", encoding="utf-8") as f:
        f.write("# This file was automatically generated\n\n")
        f.write("years = {\n")
        for k, v in years.items():
            f.write(f"        {k}: \"{v}\",\n")
        f.write("}\n\n")
        
        f.write("specialties = {\n")
        for year_num, year_specs in specialties.items():
            f.write(f"    \"{year_num}\":{{\n")
            for k, v in year_specs.items():
                f.write(f"        {k}: \"{v}\",\n")
            f.write("    },\n\n")
        f.write("}\n\n")
        
        f.write("group_list = {\n")
        for year_num, year_groups in group_list.items():
            f.write(f"    \"{year_num}\":{{\n\n")
            for spec_key, spec_groups in year_groups.items():
                f.write(f"        \"{spec_key}\": {{\n")
                for k, v in spec_groups.items():
                    f.write(f"            b\"{k.decode().replace('-', '')}\": \"{v}\",\n")
                f.write("        },\n\n")
            f.write("    },\n\n")
        f.write("}\n")
    send_logs("Dynamic group lists have been written to dynamic_group_lists.py", 'info')
    return years, specialties, group_list