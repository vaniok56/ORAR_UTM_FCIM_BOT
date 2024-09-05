# orar_utm_fcim_bot version 0.4
### changelog:
# used blackbox to speed up getMergedCellVal. Result: literally over x200 performance boost
# added the ability to track classes that are not splited by odd/even
# removed some dupe code
# the orer.xlsx is now full, not just a part
# other minor improvements

from telethon import TelegramClient, events, types
from telethon.tl.custom import Button

import configparser # read
import datetime
import openpyxl # excel read library
import datetime

import pandas as pd
import numpy as np
#import cProfile

#### Access credentials
config = configparser.ConfigParser() # Define the method to read the configuration file
config.read('config.ini') # read config.ini file

api_id = config.get('default','api_id') # get the api id
api_hash = config.get('default','api_hash') # get the api hash
BOT_TOKEN = config.get('default','BOT_TOKEN') # get the bot token

# Create the client and the session called session_master.
client = TelegramClient('sessions/session_master', api_id, api_hash).start(bot_token=BOT_TOKEN)

#classes that are not splited by odd/even
not_dual = np.array([4, 11, 24, 25, 38, 49, 50, 51])

#hours
hours =   [
    ["8:00 - 9:30"],
    ["9:45 - 11:15"],
    ["11:30 - 13:00"],
    ["13:30 - 15:00"],
    ["15:15 - 16:45"],
    ["17:00 - 18:30"],
    ["18:45 - 20:15"]
]

#week days
week_days = {
    0 : "Luni",
    1 : "Marţi",
    2 : "Miercuri",
    3 : "Joi",
    4 : "Vineri",
    5 : "Sambata",
    6 : "Duminica"
}

#group list
group_list = {
    b"ti241": "  TI-241  ",
    b"ti242": "  TI-242  ",
    b"ti243": "  TI-243  ",
    b"ti244": "  TI-244  ",
    b"ti245": "  TI-245  ",
    b"ti246": "  TI-246  ",
    b"ti247": "  TI-247  ",
    b"ti248": "  TI-248  ",
    b"fi241": "  FI-241  ",
    b"si243": "  SI-243  "
}

#keyboard buttons
bot_kb = [
        Button.text('Orarul de azi 📅', resize=True),
        Button.text('Orarul de maine 📅', resize=True),
        Button.text('Orarul saptamainii 🗓️', resize=True),
        Button.text('Orele ⏰', resize=True),
    ]

wb = openpyxl.load_workbook('orar.xlsx', data_only=True)
schedule = wb["Table 2"]

df = pd.read_csv('BD.csv') #DB

is_even = datetime.datetime.today().isocalendar().week % 2
cur_group = "TI-241" #initialise current group
groups = [schedule.cell(row=1,column=i).value for i in range(3,36)] #group list

#/start
@client.on(events.NewMessage(pattern="/(?i)start")) 
async def startt(event):
    global df
    sender = await event.get_sender()
    SENDER = sender.id
    text = "Salut!\nIn primul rand alege grupa - /alege_grupa \nPentru a afisa toate comenzile - /help"
    buttons_in_row = 2
    #bot_kb.append(types.KeyboardButtonSimpleWebView("Orar pdf🥱", "https://fcim.utm.md/procesul-de-studii/orar/#toggle-id-2-closed",))
    button_rows = button_grid(bot_kb, buttons_in_row)
    await client.send_message(SENDER, text, parse_mode="HTML", buttons=button_rows)

    #add the user to users
    if "U"+str(SENDER) not in "U"+str(df['SENDER'].to_list()):
        data =  {'SENDER' : ["U"+str(SENDER)],
                 'group' : [""]}
        new_dat = pd.DataFrame(data)
        df = pd.concat([df, new_dat])
        df.to_csv('BD.csv', encoding='utf-8', index=False)

#/help
@client.on(events.NewMessage(pattern='/(?i)help')) 
async def helpp(event):
    sender = await event.get_sender()
    SENDER = sender.id
    text  = "/start - start\n"
    text  = "/help - toate comenzile\n"
    text += "/azi - orarul de azi\n"
    text += "/maine - orarul de maine\n"
    text += "/ore - orarul orelor(perechi + pauze)\n"
    text += "/alege_grupa - alegerea grupei\n"
    text += "/sapt_curenta - orar pe saptamana curenta\n"
    text += "/sapt_viitoare - orar pe saptamana viitoare\n"
    #text += "/noti - NOT IMPLEMENTED notificari on/off\n"
    #text += "/alege - NOT IMPLEMENTED orar pe o zi concreta cu butoane\n"
    #text += "/sesiuni - NOT IMPLEMENTED orarul sesiunilor\n"
    await client.send_message(SENDER, text, parse_mode="HTML")

#/hours
@client.on(events.NewMessage(pattern='/(?i)ore|Orele ⏰')) 
async def oree(event):
    sender = await event.get_sender()
    SENDER = sender.id
    text = "Graficul de ore:\n"
    for i in range(0, 7):
        text += "\nPerechea: #" + str(i+1) + "\nOra : " + ''.join(hours[i]) + "\n"
        if i == 2 :
            text += "Pauza : " + "30 min\n"
        else:
            text += "Pauza : " + "15 min\n"
    await client.send_message(SENDER, text, parse_mode="HTML")

#/maine
@client.on(events.NewMessage(pattern='/(?i)maine|Orarul de maine 📅')) 
async def mainee(event):
    global df, cur_group
    week_day = int((datetime.datetime.today() + datetime.timedelta(days=1)).weekday()) #weekday tomorrow(0-6)
    sender = await event.get_sender()
    SENDER = sender.id
    try:
        #get the user's selected group
        csv_gr = list(df.loc[df['SENDER'] == "U"+str(SENDER), 'group'])[0]
        cur_group = csv_gr
        if cur_group == "":
            raise ValueError('no gr')
        else: 
            #send the schedule
            text = "\nGrupa - " + cur_group + "\nOrarul de maine(" + week_days[week_day] +"):" + print_day(week_day)
            await client.send_message(SENDER, text, parse_mode="HTML")
    except Exception as error:
        print("An exception occurred:", error)
        await client.send_message(SENDER, "Inca nu ai ales grupa.\n/alege_grupa", parse_mode="HTML")
    
#/azi
@client.on(events.NewMessage(pattern='/(?i)azi|Orarul de azi 📅')) 
async def azii(event):
    global df, cur_group
    week_day = int(datetime.datetime.today().weekday()) #weekday today(0-6)
    sender = await event.get_sender()
    SENDER = sender.id
    try:
        csv_gr = list(df.loc[df['SENDER'] == "U"+str(SENDER), 'group'])[0]
        cur_group = csv_gr
        if cur_group == "":
            raise ValueError('no gr')
        else: 
            text = "\nGrupa - " + cur_group + "\nOrarul de azi(" + week_days[week_day] +"):" + print_day(week_day)
            await client.send_message(SENDER, text, parse_mode="HTML")
    except Exception as error:
        print("An exception occurred:", error)
        print(df)
        await client.send_message(SENDER, "Inca nu ai ales grupa.\n/alege_grupa", parse_mode="HTML")
    

#/sapt_cur
@client.on(events.NewMessage(pattern='/(?i)sapt_curenta|Orarul saptamainii 🗓️')) 
async def sapt_curr(event):
    global df, cur_group, is_even
    sender = await event.get_sender()
    SENDER = sender.id
    try:
        csv_gr = list(df.loc[df['SENDER'] == "U"+str(SENDER), 'group'])[0]
        cur_group = csv_gr
        if cur_group == "":
            raise ValueError('no gr')
        else: 
            text = "\nGrupa - " + cur_group + "\nOrarul pe saptamana aceasta:" + print_sapt(is_even)
            await client.send_message(SENDER, text, parse_mode="HTML")
    except Exception as error:
        print("An exception occurred:", error)
        await client.send_message(SENDER, "Inca nu ai ales grupa.\n/alege_grupa", parse_mode="HTML")

#/sapt_viit
@client.on(events.NewMessage(pattern='/(?i)sapt_viitoare')) 
async def sapt_viit(event):
    global df, cur_group, is_even
    is_even = not is_even
    sender = await event.get_sender()
    SENDER = sender.id

    try:
        csv_gr = list(df.loc[df['SENDER'] == "U"+str(SENDER), 'group'])[0]
        cur_group = csv_gr
        if cur_group == "":
            raise ValueError('no gr')
        else: 
            text = "\nGrupa - " + cur_group + "\nOrarul pe saptamana viitoare:" + print_sapt(is_even)
            await client.send_message(SENDER, text, parse_mode="HTML")
    except Exception as error:
        print("An exception occurred:", error)
        await client.send_message(SENDER, "Inca nu ai ales grupa.\n/alege_grupa", parse_mode="HTML")

def button_grid(buttons, butoane_rand):
    return [buttons[i:i + butoane_rand] for i in range(0, len(buttons), butoane_rand)]

#/alege_grupa
@client.on(events.NewMessage(pattern='/(?i)alege_grupa')) 
async def alege_grupaa(event):
    sender = await event.get_sender()
    SENDER = sender.id
    text = "Alege grupa"
    buttons = [Button.inline(group, data=data) for data, group in group_list.items()]
    button_per_r = 4
    button_rows = button_grid(buttons, button_per_r)
    await client.send_message(SENDER, text, parse_mode="HTML", buttons=button_rows)

#button click event handle
@client.on(events.CallbackQuery())
async def callback(event):
    global df, cur_group
    sender = await event.get_sender()
    SENDER = sender.id
    cur_group = group_list.get(event.data).replace(" ", "")
    if cur_group:
        #if user is not in list, add it
        if "U"+str(SENDER) not in "U"+str(df['SENDER'].to_list()):
            data =  {'SENDER' : ["U"+str(SENDER)],
                     'group' : [""]}
            new_dat = pd.DataFrame(data)
            df = pd.concat([df, new_dat]) 
            df.to_csv('BD.csv', encoding='utf-8', index=False)

        #updates the current group
        df.loc[df['SENDER'] == "U"+str(SENDER), 'group'] = cur_group #send cur_group to df
        df.to_csv('BD.csv', encoding='utf-8', index=False) #save df

        print(df.loc[df['SENDER'] == "U"+str(SENDER)])
        
        sender = await event.get_sender()
        SENDER = sender.id
        text = f"Grupa ta este: {cur_group}"
        await event.answer('Grupa a fost selectata!')
        await client.send_message(SENDER, text, parse_mode="HTML")
        
#get value from a cell even if it's a merged cell
merged_cell_ranges = {}  # cache merged cell ranges

def get_merged_cell_ranges(sheet):
    global merged_cell_ranges
    if sheet not in merged_cell_ranges:
        merged_cell_ranges[sheet] = {(r.min_row, r.min_col): r for r in sheet.merged_cells.ranges}
    return merged_cell_ranges[sheet]

def getMergedCellVal(sheet, cell):
    merged_ranges = get_merged_cell_ranges(sheet)
    row, col = cell.row, cell.column
    for (r_min, c_min), rng in merged_ranges.items():
        if r_min <= row <= rng.max_row and c_min <= col <= rng.max_col:
            return sheet.cell(rng.min_row, rng.min_col).value
    return cell.value

#get daily schedule
def print_day(week_day) :
    global schedule, is_even, groups
    col_gr = groups.index(cur_group) + 3 #column with the selected group
    row_start = 2 + (14 * week_day) #first course row

    daily = print_daily(schedule, row_start, is_even, col_gr)

    return daily

def print_daily(schedule, row_start, is_even, col_gr):
    day_sch = []
    seen = set() 

    #rowstart depending on not_dual
    row_start -= sum(1 for i in range(1, row_start) if np.isin(i, not_dual) and i != 51)
    if is_even == True:
        row_start+=1
    match_is_even = not is_even

    #get the hours list
    orele = {i: getMergedCellVal(schedule, schedule.cell(row=i, column=2)) for i in range(row_start, row_start + 13)}

    #export all courses into a dataframe
    for i in range(row_start, row_start + 13):
        ora = orele[i] #get curent hour
        is_not_dual = False
        match_is_even = not match_is_even

        #check if current row is in not_dual
        if np.isin(i, not_dual):
            is_not_dual = True
            match_is_even = not match_is_even
        #jump to next iteration if already seen this hour course
        if ora in seen:
            continue
        #Add the course if even/odd or a row was skiped
        if match_is_even == is_even or is_not_dual == True :
            seen.add(ora)
            day_sch.append(getMergedCellVal(schedule, schedule.cell(row=i, column=col_gr)))
        else: 
            continue #jump to next iteration if odd/even

    #modifying for beter visibility
    for i in range(len(day_sch)):
        day_sch[i] = "<b>" + str(day_sch[i]) + "</b>"
        if "None" in str(day_sch[i]):
            day_sch[i] = ""
        else: day_sch[i] = "\nPerechea: #" + str(i+1) + "\n" + str(day_sch[i]) + "\nOra : " + ''.join(hours[i]) + "\n"
    
    #dataframe to string
    day_sch = "\n" + "".join(str(element) for element in day_sch)
    
    return day_sch

#get weekly schedule
def print_sapt(is_even) :
    global schedule, groups
    col_gr = groups.index(cur_group) + 3 #column with the selected group
    row_start = 2 #first course row

    week_sch = ""
    for j in range(1, 6):
        daily = print_daily(schedule, row_start, is_even, col_gr)

        #do not print an empty weekday
        if str(daily) == "":
            week_sch += "\n\n"
        else: 
            week_sch += "\n\n&emsp;&emsp;&emsp;&emsp;<b>" + week_days[j-1] + ":</b>" + "\n" + daily
        
        row_start += 14
    return week_sch

### MAIN
if __name__ == '__main__':
    print("Bot Started!")
    #cProfile.run('print_sapt(is_even)', sort='tottime')
    client.run_until_disconnected()
