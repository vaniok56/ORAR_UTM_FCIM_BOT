import openpyxl
import pandas as pd
import numpy as np
import datetime

ore =   [["8:00 - 9:30"],
         ["9:45 - 11:15"],
        ["11:30 - 13:00"],
        ["13:30 - 15:00"],
        ["15:15 - 16:45"],
        ["17:00 - 18:30"],
        ["18:45 - 20:15"]]


curr_date = datetime.datetime.today()
grupa = 'TI-243'
week_day = curr_date.weekday()

def getMergedCellVal(sheet, cell):
    rng = [s for s in sheet.merged_cells.ranges if cell.coordinate in s]
    return sheet.cell(rng[0].min_row, rng[0].min_col).value if len(rng)!=0 else cell.value

def print_azi() :
    # excel open
    wb = openpyxl.load_workbook('orarf test.xlsx', data_only=True)
    orar = wb["Table 2"]
    curr_date = datetime.datetime.today() # data de azi
    week_day = curr_date.weekday() #ziua saptamanii in nr (0-6)
    para = int(curr_date.isocalendar().week%2)
    
    # para = True # remove
    grupe = [orar.cell(row=1,column=i).value for i in range(3,37)] #lista toate grupe
    col_gr = grupe.index(grupa) + 3 #coloana cu gupa selectata anterior
    print(orar['A1'].value)
    row_start = 2 + (14*week_day) + int(para)

    row_start = 3 # remove

    orar_azi = []
    vazute = set() #
    for i in range(row_start, row_start + 13):
        ora = getMergedCellVal(orar, orar.cell(row=i, column=2)) #ora
        if ora in vazute:
            continue  # skip
        vazute.add(ora)
        orar_azi.append(getMergedCellVal(orar, orar.cell(row=i, column=col_gr)))

    for i in range(len(orar_azi)):
        orar_azi[i] = "<b>" + str(orar_azi[i]) + "</b>"
        if "None" in str(orar_azi[i]):
            orar_azi[i] = "<b>---</b>"
        orar_azi[i] = "Perechea: #" + str(i+1) + "\n" + str(orar_azi[i]) + "\nOra : " + ''.join(ore[i])
    orar_azi = "\n\n".join(str(element) for element in orar_azi)
    wb.close()
    return orar_azi


print("orar:\n\n" + print_azi())